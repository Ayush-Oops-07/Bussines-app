"""
backend/app/api/endpoints/reports.py — Reports API router.

Exposes endpoints for Monthly Sales Report and Customer Ledger.
Generates server-side PDFs using ReportLab and Excels using openpyxl.
"""

import io
import uuid
import calendar
from datetime import date, datetime
from typing import Optional
from decimal import Decimal

import os
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

# Backend imports
from app.core.dependencies import get_db, get_current_user
from app.models.models import User, Party, LedgerEntry, InvoiceAdjustment
from app.repositories import party_repository, invoice_repository, return_repository
from app.services import ledger_service

# ReportLab imports for PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# openpyxl imports for Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ── HELPERS ───────────────────────────────────────────────────────────────────

def resolve_dates(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None
) -> tuple[date, date]:
    """Resolves inputs to a start and end date range."""
    if month is not None and year is not None:
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, 1), date(year, month, last_day)
    
    if not from_date and not to_date:
        today = date.today()
        last_day = calendar.monthrange(today.year, today.month)[1]
        return date(today.year, today.month, 1), date(today.year, today.month, last_day)
    
    resolved_from = from_date or date(2000, 1, 1)
    resolved_to = to_date or date(2099, 12, 31)
    return resolved_from, resolved_to


def format_inr(number) -> str:
    """Format decimal/float as Indian Currency format (e.g. ₹12,34,567.89)."""
    try:
        if number is None:
            return "₹0.00"
        val = float(number)
        is_negative = val < 0
        val = abs(val)
        
        s = f"{val:.2f}"
        parts = s.split(".")
        num = parts[0]
        dec = parts[1]
        
        if len(num) <= 3:
            res = num
        else:
            last_three = num[-3:]
            remaining = num[:-3]
            groups = []
            while remaining:
                groups.append(remaining[-2:])
                remaining = remaining[:-2]
            groups.reverse()
            res = ",".join(groups) + "," + last_three
        
        prefix = "-" if is_negative else ""
        return prefix + "₹" + res + "." + dec
    except Exception:
        return f"₹{number}"


# ── REPORTLAB CUSTOM CANVAS FOR A4 PAGE NUMBERING ────────────────────────────

class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to calculate total page count and draw
    professional headers and footers with page numbers.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_header_footer(num_pages)
            super().showPage()
        super().save()

    def draw_header_footer(self, page_count):
        self.saveState()
        
        # Color codes: primary dark (#1e293b), border (#e2e8f0)
        slate_600 = colors.HexColor("#475569")
        slate_300 = colors.HexColor("#cbd5e1")
        
        # Header (Top margin ends at 787.89)
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(slate_600)
        self.drawString(54, 800, "SANDEEP TRADERS")
        
        report_title = getattr(self, "_report_title", "BUSINESS REPORT")
        self.drawRightString(541, 800, report_title.upper())
        
        self.setStrokeColor(slate_300)
        self.setLineWidth(0.5)
        self.line(54, 792, 541, 792)
        
        # Footer (Bottom margin starts at 54)
        self.setFont("Helvetica", 8)
        self.drawString(54, 35, f"Generated Date: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
        
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(541, 35, page_text)
        
        self.line(54, 47, 541, 47)
        self.restoreState()


def make_numbered_canvas(report_title: str):
    class CustomNumberedCanvas(NumberedCanvas):
        _report_title = report_title
    return CustomNumberedCanvas


# ── GET MONTHLY SALES REPORT DATA ─────────────────────────────────────────────

async def fetch_monthly_sales_data(
    db: AsyncSession,
    resolved_from: date,
    resolved_to: date
) -> dict:
    """Helper to query sales invoice ledger entries and compute metrics."""
    # Subquery to sum adjustments for each invoice
    adj_sub = (
        select(
            InvoiceAdjustment.invoice_id,
            func.coalesce(func.sum(InvoiceAdjustment.amount), 0).label("paid_sum")
        )
        .where(InvoiceAdjustment.is_deleted == False)
        .group_by(InvoiceAdjustment.invoice_id)
        .subquery()
    )

    # Main query for customer sale ledger entries
    query = (
        select(
            LedgerEntry,
            Party,
            func.coalesce(adj_sub.c.paid_sum, 0).label("paid_amount")
        )
        .join(Party, Party.id == LedgerEntry.party_id)
        .outerjoin(adj_sub, adj_sub.c.invoice_id == LedgerEntry.invoice_id)
        .where(
            LedgerEntry.entry_type == "sale",
            LedgerEntry.is_deleted == False,
            Party.party_type == "customer",
            LedgerEntry.entry_date >= resolved_from,
            LedgerEntry.entry_date <= resolved_to
        )
        .order_by(LedgerEntry.entry_date.asc(), LedgerEntry.invoice_number.asc())
    )

    res = await db.execute(query)
    rows = res.all()

    invoices = []
    total_sales = Decimal("0.00")
    total_paid = Decimal("0.00")
    total_outstanding = Decimal("0.00")

    for entry, party, paid_val in rows:
        invoice_total = Decimal(str(entry.debit))
        paid_amount = Decimal(str(paid_val))
        
        # Calculate outstanding safely
        outstanding = invoice_total - paid_amount
        if outstanding < 0:
            outstanding = Decimal("0.00")
            
        # Determine status
        if paid_amount >= invoice_total:
            status = "Paid"
        elif paid_amount > 0:
            status = "Partially Paid"
        else:
            status = "Unpaid"

        invoices.append({
            "invoice_number": entry.invoice_number or "",
            "invoice_date": entry.entry_date.isoformat() if entry.entry_date else "",
            "customer_name": party.name,
            "mobile": party.mobile or "",
            "total_amount": float(invoice_total),
            "payment_status": status,
            "outstanding": float(outstanding)
        })

        total_sales += invoice_total
        total_paid += paid_amount
        total_outstanding += outstanding

    return {
        "total_sales": float(total_sales),
        "total_paid": float(total_paid),
        "total_outstanding": float(total_outstanding),
        "num_invoices": len(invoices),
        "invoices": invoices
    }


# ── API ENDPOINTS ─────────────────────────────────────────────────────────────

@router.get("/monthly-sales")
async def get_monthly_sales_report(
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    data = await fetch_monthly_sales_data(db, resolved_from, resolved_to)
    data["from_date"] = resolved_from.isoformat()
    data["to_date"] = resolved_to.isoformat()
    return data



# ── PDF GENERATOR: MONTHLY SALES ──────────────────────────────────────────────

@router.get("/monthly-sales/pdf")
async def get_monthly_sales_pdf(
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    data = await fetch_monthly_sales_data(db, resolved_from, resolved_to)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=15
    )
    normal_style = ParagraphStyle(
        'DocNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#475569")
    )
    bold_style = ParagraphStyle(
        'DocBold',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    elements = []
    
    # Header block
    elements.append(Paragraph("SANDEEP TRADERS", title_style))
    elements.append(Paragraph("Monthly Sales Report", ParagraphStyle('Sub', parent=title_style, fontSize=12, leading=15, textColor=colors.HexColor("#64748b"))))
    
    date_str = f"Period: {resolved_from.strftime('%d-%b-%Y')} to {resolved_to.strftime('%d-%b-%Y')}"
    if month is not None and year is not None:
        date_str = f"Report for: {calendar.month_name[month]} {year}"
    elements.append(Paragraph(date_str, normal_style))
    elements.append(Spacer(1, 15))
    
    # Table data
    table_data = [
        ["Invoice No", "Date", "Customer Name", "Mobile", "Total Amount", "Status", "Outstanding"]
    ]
    
    for inv in data["invoices"]:
        table_data.append([
            inv["invoice_number"],
            inv["invoice_date"],
            inv["customer_name"],
            inv["mobile"],
            format_inr(inv["total_amount"]),
            inv["payment_status"],
            format_inr(inv["outstanding"])
        ])
        
    # Table definition and sizing (width limit = 487pt)
    # Total widths: 65 + 65 + 110 + 70 + 65 + 60 + 52 = 487
    col_widths = [65, 65, 110, 70, 65, 60, 52]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),  # Total Amount
        ('ALIGN', (6, 0), (6, -1), 'RIGHT'),  # Outstanding
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ])
    
    # Alternating rows
    for i in range(1, len(table_data)):
        bg = colors.HexColor("#f8fafc") if i % 2 == 1 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        
    table.setStyle(t_style)
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # Bottom Summary block
    summary_data = [
        [Paragraph("<b>Number of Invoices:</b>", normal_style), Paragraph(str(data["num_invoices"]), normal_style),
         Paragraph("<b>Total Sales:</b>", normal_style), Paragraph(format_inr(data["total_sales"]), bold_style)],
        ["", "", Paragraph("<b>Total Paid:</b>", normal_style), Paragraph(format_inr(data["total_paid"]), bold_style)],
        ["", "", Paragraph("<b>Total Outstanding:</b>", normal_style), Paragraph(format_inr(data["total_outstanding"]), bold_style)]
    ]
    
    summary_table = Table(summary_data, colWidths=[120, 80, 150, 137])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEABOVE', (2, 0), (3, 0), 1, colors.HexColor("#0f172a")),
    ]))
    elements.append(summary_table)
    
    doc.build(elements, canvasmaker=make_numbered_canvas("Monthly Sales Report"))
    
    buffer.seek(0)
    filename = f"monthly_sales_{resolved_from.isoformat()}_to_{resolved_to.isoformat()}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── PDF GENERATOR: CUSTOMER LEDGER ────────────────────────────────────────────

@router.get("/customer-ledger/pdf")
async def get_customer_ledger_pdf(
    customer_id: uuid.UUID,
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    
    party = await party_repository.get_by_id(db, customer_id)
    if not party:
        raise HTTPException(status_code=404, detail="Customer not found")
        
    ledger_data = await ledger_service.get_ledger(
        db, party_id=customer_id, from_date=resolved_from, to_date=resolved_to
    )
    
    entries = ledger_data.get("entries", [])
    opening_balance = Decimal(str(ledger_data.get("opening_balance", 0.0)))
    
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    for e in entries:
        total_debit += Decimal(str(e.get("debit", 0.0)))
        total_credit += Decimal(str(e.get("credit", 0.0)))
    
    final_balance = opening_balance + total_debit - total_credit
    current_outstanding = final_balance if final_balance > 0 else Decimal("0.00")
    advance_balance = abs(final_balance) if final_balance < 0 else Decimal("0.00")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=15
    )
    normal_style = ParagraphStyle(
        'DocNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#475569")
    )
    bold_style = ParagraphStyle(
        'DocBold',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    elements = []
    
    # Document header
    elements.append(Paragraph("SANDEEP TRADERS", title_style))
    elements.append(Paragraph("Customer Ledger Statement", ParagraphStyle('Sub', parent=title_style, fontSize=12, leading=15, textColor=colors.HexColor("#64748b"))))
    elements.append(Spacer(1, 10))
    
    # Customer Details Block
    cust_details = [
        [Paragraph(f"<b>Customer Name:</b> {party.name}", normal_style),
         Paragraph(f"<b>Period:</b> {resolved_from.strftime('%d-%b-%Y')} to {resolved_to.strftime('%d-%b-%Y')}", normal_style)],
        [Paragraph(f"<b>Mobile:</b> {party.mobile or 'N/A'}", normal_style),
         Paragraph(f"<b>Opening Balance:</b> {format_inr(opening_balance)}", normal_style)],
        [Paragraph(f"<b>Address:</b> {party.address or 'N/A'}", normal_style),
         Paragraph(f"<b>Closing Balance:</b> {format_inr(final_balance)}", normal_style)]
    ]
    cust_table = Table(cust_details, colWidths=[240, 247])
    cust_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    elements.append(cust_table)
    elements.append(Spacer(1, 15))
    
    # Ledger Table
    table_data = [
        ["Date", "Particulars", "Debit (DR)", "Credit (CR)", "Running Balance"]
    ]
    
    for e in entries:
        particulars = e.get("particulars", "")
        if e.get("notes"):
            particulars += f" ({e.get('notes')})"
        table_data.append([
            e.get("entry_date", ""),
            particulars,
            format_inr(e.get("debit", 0.0)) if float(e.get("debit", 0.0)) > 0 else "—",
            format_inr(e.get("credit", 0.0)) if float(e.get("credit", 0.0)) > 0 else "—",
            format_inr(e.get("running_balance", 0.0))
        ])
        
    # Total widths: 70 + 187 + 75 + 75 + 80 = 487
    col_widths = [70, 187, 75, 75, 80]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),  # Debit
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),  # Credit
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),  # Running Balance
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ])
    
    for i in range(1, len(table_data)):
        bg = colors.HexColor("#f8fafc") if i % 2 == 1 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        
    table.setStyle(t_style)
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # Bottom Summary block
    summary_data = [
        [Paragraph("<b>Total Debit (DR):</b>", normal_style), Paragraph(format_inr(total_debit), bold_style),
         Paragraph("<b>Current Outstanding:</b>", normal_style), Paragraph(format_inr(current_outstanding), bold_style)],
        [Paragraph("<b>Total Credit (CR):</b>", normal_style), Paragraph(format_inr(total_credit), bold_style),
         Paragraph("<b>Advance Balance:</b>", normal_style), Paragraph(format_inr(advance_balance), bold_style)]
    ]
    
    summary_table = Table(summary_data, colWidths=[120, 120, 120, 127])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEABOVE', (0, 0), (1, 0), 1, colors.HexColor("#0f172a")),
        ('LINEABOVE', (2, 0), (3, 0), 1, colors.HexColor("#0f172a")),
    ]))
    elements.append(summary_table)
    
    doc.build(elements, canvasmaker=make_numbered_canvas("Customer Ledger Statement"))
    
    buffer.seek(0)
    filename = f"ledger_{party.name.replace(' ', '_')}_{resolved_from.isoformat()}_to_{resolved_to.isoformat()}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── EXCEL GENERATOR: MONTHLY SALES ────────────────────────────────────────────

@router.get("/monthly-sales/excel")
async def get_monthly_sales_excel(
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    data = await fetch_monthly_sales_data(db, resolved_from, resolved_to)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Sales Report"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_accent = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='0F172A'),
        bottom=Side(style='double', color='0F172A')
    )
    
    # Header block
    ws["A1"] = "SANDEEP TRADERS"
    ws["A1"].font = font_title
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = "Monthly Sales Report"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True, color="475569")
    
    date_str = f"Period: {resolved_from.strftime('%d-%b-%Y')} to {resolved_to.strftime('%d-%b-%Y')}"
    if month is not None and year is not None:
        date_str = f"Report for: {calendar.month_name[month]} {year}"
    ws["A3"] = date_str
    ws["A3"].font = font_subtitle
    
    # Table headers
    headers = ["Invoice No", "Invoice Date", "Customer Name", "Mobile", "Total Amount", "Payment Status", "Outstanding"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left" if col_idx != 5 and col_idx != 7 else "right", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[5].height = 24
    
    # Rows
    current_row = 6
    for inv in data["invoices"]:
        ws.cell(row=current_row, column=1, value=inv["invoice_number"]).alignment = Alignment(horizontal="left")
        ws.cell(row=current_row, column=2, value=inv["invoice_date"]).alignment = Alignment(horizontal="left")
        ws.cell(row=current_row, column=3, value=inv["customer_name"]).alignment = Alignment(horizontal="left")
        ws.cell(row=current_row, column=4, value=inv["mobile"]).alignment = Alignment(horizontal="left")
        
        c_tot = ws.cell(row=current_row, column=5, value=inv["total_amount"])
        c_tot.number_format = '"₹"#,##,##0.00'
        c_tot.alignment = Alignment(horizontal="right")
        
        ws.cell(row=current_row, column=6, value=inv["payment_status"]).alignment = Alignment(horizontal="left")
        
        c_out = ws.cell(row=current_row, column=7, value=inv["outstanding"])
        c_out.number_format = '"₹"#,##,##0.00'
        c_out.alignment = Alignment(horizontal="right")
        
        for col_idx in range(1, 8):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_body
            c.border = thin_border
            if current_row % 2 == 1:
                c.fill = fill_accent
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Totals Summary row
    ws.cell(row=current_row, column=1, value="Total").font = font_bold
    ws.cell(row=current_row, column=1).border = double_bottom_border
    ws.cell(row=current_row, column=2, value=f"Invoices: {data['num_invoices']}").font = font_bold
    ws.cell(row=current_row, column=2).border = double_bottom_border
    
    for c in range(3, 5):
        ws.cell(row=current_row, column=c).border = double_bottom_border
        
    t_sales_cell = ws.cell(row=current_row, column=5, value=data["total_sales"])
    t_sales_cell.font = font_bold
    t_sales_cell.number_format = '"₹"#,##,##0.00'
    t_sales_cell.alignment = Alignment(horizontal="right")
    t_sales_cell.border = double_bottom_border
    
    t_paid_cell = ws.cell(row=current_row, column=6, value=f"Paid: {format_inr(data['total_paid'])}").font = font_bold
    t_paid_cell.border = double_bottom_border
    
    t_out_cell = ws.cell(row=current_row, column=7, value=data["total_outstanding"])
    t_out_cell.font = font_bold
    t_out_cell.number_format = '"₹"#,##,##0.00'
    t_out_cell.alignment = Alignment(horizontal="right")
    t_out_cell.border = double_bottom_border
    
    ws.row_dimensions[current_row].height = 22
    
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Scan header and body rows only for size computation (rows 5 to current_row)
        for cell in col[4:current_row]:
            if cell.value:
                # Add formatting padding
                val_str = str(cell.value)
                if isinstance(cell.value, float) or isinstance(cell.value, Decimal):
                    val_str = f"₹{cell.value:,.2f}"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"monthly_sales_{resolved_from.isoformat()}_to_{resolved_to.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── EXCEL GENERATOR: CUSTOMER LEDGER ──────────────────────────────────────────

@router.get("/customer-ledger/excel")
async def get_customer_ledger_excel(
    customer_id: uuid.UUID,
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    
    party = await party_repository.get_by_id(db, customer_id)
    if not party:
        raise HTTPException(status_code=404, detail="Customer not found")
        
    ledger_data = await ledger_service.get_ledger(
        db, party_id=customer_id, from_date=resolved_from, to_date=resolved_to
    )
    
    entries = ledger_data.get("entries", [])
    opening_balance = Decimal(str(ledger_data.get("opening_balance", 0.0)))
    
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    for e in entries:
        total_debit += Decimal(str(e.get("debit", 0.0)))
        total_credit += Decimal(str(e.get("credit", 0.0)))
    
    final_balance = opening_balance + total_debit - total_credit
    current_outstanding = final_balance if final_balance > 0 else Decimal("0.00")
    advance_balance = abs(final_balance) if final_balance < 0 else Decimal("0.00")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_accent = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_info = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='0F172A'),
        bottom=Side(style='double', color='0F172A')
    )
    
    # Header block
    ws["A1"] = "SANDEEP TRADERS"
    ws["A1"].font = font_title
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = "Customer Ledger Statement"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True, color="475569")
    
    date_str = f"Period: {resolved_from.strftime('%d-%b-%Y')} to {resolved_to.strftime('%d-%b-%Y')}"
    ws["A3"] = date_str
    ws["A3"].font = font_subtitle
    
    # Info Block (A5 to E7)
    ws["A5"] = "Customer Name:"
    ws["A5"].font = font_bold
    ws["B5"] = party.name
    ws["B5"].font = font_body
    
    ws["D5"] = "Period Range:"
    ws["D5"].font = font_bold
    ws["E5"] = f"{resolved_from.isoformat()} to {resolved_to.isoformat()}"
    ws["E5"].font = font_body
    
    ws["A6"] = "Mobile Number:"
    ws["A6"].font = font_bold
    ws["B6"] = party.mobile or "N/A"
    ws["B6"].font = font_body
    
    ws["D6"] = "Opening Balance:"
    ws["D6"].font = font_bold
    ws["E6"] = float(opening_balance)
    ws["E6"].font = font_body
    ws["E6"].number_format = '"₹"#,##,##0.00'
    
    ws["A7"] = "Address Details:"
    ws["A7"].font = font_bold
    ws["B7"] = party.address or "N/A"
    ws["B7"].font = font_body
    
    ws["D7"] = "Closing Balance:"
    ws["D7"].font = font_bold
    ws["E7"] = float(final_balance)
    ws["E7"].font = font_body
    ws["E7"].number_format = '"₹"#,##,##0.00'
    
    # Apply borders & background to Info Block
    for r in range(5, 8):
        for col_idx in range(1, 6):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = thin_border
            cell.fill = fill_info
            
    # Table headers
    headers = ["Date", "Particular", "Invoice Link", "Return Link", "Debit (DR)", "Credit (CR)", "Running Balance"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left" if col_idx <= 4 else "right", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[9].height = 24
    
    created_sheets = set()
    
    frontend_url = os.getenv("FRONTEND_URL", "http://localhost:3000").rstrip("/")
    
    # Data Rows
    current_row = 10
    for e in entries:
        ws.cell(row=current_row, column=1, value=e.get("entry_date", "")).alignment = Alignment(horizontal="left")
        
        particulars = e.get("particulars", "")
        if e.get("notes"):
            particulars += f" ({e.get('notes')})"
        ws.cell(row=current_row, column=2, value=particulars).alignment = Alignment(horizontal="left")
        
        # Return Link
        pr_id = e.get("purchase_return_id")
        c_ret_link = ws.cell(row=current_row, column=4)
        pr_ref_invoice_id = None
        if pr_id:
            try:
                pr = await return_repository.get_by_id(db, uuid.UUID(pr_id), load_items=True)
                if pr:
                    pr_ref_invoice_id = pr.reference_invoice_id
                    pr_sheet_title = f"Ret {pr.return_number}"[:31]
                    if pr_id not in created_sheets:
                        pr_ws = wb.create_sheet(title=pr_sheet_title)
                        pr_ws["A1"] = f"Return: {pr.return_number}"
                        pr_ws["A2"] = f"Date: {pr.return_date}"
                        pr_ws["A3"] = f"Total Amount: ₹{float(pr.total_amount):.2f}"
                        pr_ws["A1"].font = font_bold
                        pr_ws["A5"] = "Item Name"
                        pr_ws["B5"] = "Qty"
                        pr_ws["C5"] = "Rate"
                        pr_ws["D5"] = "Total"
                        for c in range(1, 5):
                            pr_ws.cell(row=5, column=c).font = font_bold
                        r = 6
                        for item in pr.items:
                            pr_ws.cell(row=r, column=1, value=item.product_name)
                            pr_ws.cell(row=r, column=2, value=float(item.quantity))
                            pr_ws.cell(row=r, column=3, value=float(item.rate))
                            pr_ws.cell(row=r, column=4, value=float(item.total))
                            r += 1
                        pr_ws.cell(row=r+2, column=1, value="Back to Ledger").hyperlink = "#'Customer Ledger'!A1"
                        pr_ws.cell(row=r+2, column=1).font = Font(color="0563C1", underline="single")
                        
                        pr_ws.column_dimensions["A"].width = 30
                        pr_ws.column_dimensions["B"].width = 15
                        pr_ws.column_dimensions["C"].width = 15
                        pr_ws.column_dimensions["D"].width = 15
                        created_sheets.add(pr_id)
                        
                    c_ret_link.value = f"View {pr.return_number}"
                    c_ret_link.hyperlink = f"#'{pr_sheet_title}'!A1"
                    c_ret_link.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
                else:
                    c_ret_link.value = "—"
            except Exception:
                c_ret_link.value = "—"
        else:
            c_ret_link.value = "—"
            c_ret_link.alignment = Alignment(horizontal="center")
            
        # Invoice Link
        invoice_id = e.get("invoice_id")
        if not invoice_id and pr_ref_invoice_id:
            invoice_id = str(pr_ref_invoice_id)
            
        c_inv_link = ws.cell(row=current_row, column=3)
        if invoice_id:
            try:
                inv = await invoice_repository.get_by_id(db, uuid.UUID(invoice_id), load_items=True)
                if inv:
                    inv_sheet_title = f"Inv {inv.invoice_number}"[:31]
                    if invoice_id not in created_sheets:
                        inv_ws = wb.create_sheet(title=inv_sheet_title)
                        inv_ws["A1"] = f"Invoice: {inv.invoice_number}"
                        inv_ws["A2"] = f"Date: {inv.invoice_date}"
                        inv_ws["A3"] = f"Total Amount: ₹{float(inv.total_amount):.2f}"
                        inv_ws["A1"].font = font_bold
                        inv_ws["A5"] = "Item Name"
                        inv_ws["B5"] = "Qty"
                        inv_ws["C5"] = "Rate"
                        inv_ws["D5"] = "Total"
                        for c in range(1, 5):
                            inv_ws.cell(row=5, column=c).font = font_bold
                        r = 6
                        for item in inv.items:
                            inv_ws.cell(row=r, column=1, value=item.product_name)
                            inv_ws.cell(row=r, column=2, value=float(item.quantity))
                            inv_ws.cell(row=r, column=3, value=float(item.rate))
                            inv_ws.cell(row=r, column=4, value=float(item.total))
                            r += 1
                        inv_ws.cell(row=r+2, column=1, value="Back to Ledger").hyperlink = "#'Customer Ledger'!A1"
                        inv_ws.cell(row=r+2, column=1).font = Font(color="0563C1", underline="single")
                        
                        inv_ws.column_dimensions["A"].width = 30
                        inv_ws.column_dimensions["B"].width = 15
                        inv_ws.column_dimensions["C"].width = 15
                        inv_ws.column_dimensions["D"].width = 15
                        created_sheets.add(invoice_id)
                    
                    c_inv_link.value = f"View {inv.invoice_number}"
                    c_inv_link.hyperlink = f"#'{inv_sheet_title}'!A1"
                    c_inv_link.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
                else:
                    c_inv_link.value = "—"
            except Exception:
                c_inv_link.value = "—"
        else:
            c_inv_link.value = "—"
            c_inv_link.alignment = Alignment(horizontal="center")

        # Debit
        deb_val = float(e.get("debit", 0.0))
        c_deb = ws.cell(row=current_row, column=5)
        if deb_val > 0:
            c_deb.value = deb_val
            c_deb.number_format = '"₹"#,##,##0.00'
        else:
            c_deb.value = "—"
        c_deb.alignment = Alignment(horizontal="right")
        
        # Credit
        cred_val = float(e.get("credit", 0.0))
        c_cred = ws.cell(row=current_row, column=6)
        if cred_val > 0:
            c_cred.value = cred_val
            c_cred.number_format = '"₹"#,##,##0.00'
        else:
            c_cred.value = "—"
        c_cred.alignment = Alignment(horizontal="right")
        
        # Running Balance
        bal_val = float(e.get("running_balance", 0.0))
        c_bal = ws.cell(row=current_row, column=7, value=bal_val)
        c_bal.number_format = '"₹"#,##,##0.00'
        c_bal.alignment = Alignment(horizontal="right")
        
        for col_idx in range(1, 8):
            c = ws.cell(row=current_row, column=col_idx)
            # Retain hyperlink font if it's the link column with an actual link
            if not ((col_idx == 3 and invoice_id) or (col_idx == 4 and pr_id)):
                c.font = font_body
            c.border = thin_border
            if current_row % 2 == 1:
                c.fill = fill_accent
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Bottom Summary rows
    # Row 1: Total Debit & Current Outstanding
    ws.cell(row=current_row, column=1, value="Total Debit:").font = font_bold
    ws.cell(row=current_row, column=1).border = double_bottom_border
    
    ws.cell(row=current_row, column=2).border = double_bottom_border
    ws.cell(row=current_row, column=3).border = double_bottom_border
    ws.cell(row=current_row, column=4).border = double_bottom_border
    
    td_cell = ws.cell(row=current_row, column=5, value=float(total_debit))
    td_cell.font = font_bold
    td_cell.number_format = '"₹"#,##,##0.00'
    td_cell.alignment = Alignment(horizontal="right")
    td_cell.border = double_bottom_border
    
    ws.cell(row=current_row, column=6, value="Current Outstanding:").font = font_bold
    ws.cell(row=current_row, column=6).border = double_bottom_border
    
    co_cell = ws.cell(row=current_row, column=7, value=float(current_outstanding))
    co_cell.font = font_bold
    co_cell.number_format = '"₹"#,##,##0.00'
    co_cell.alignment = Alignment(horizontal="right")
    co_cell.border = double_bottom_border
    ws.row_dimensions[current_row].height = 22
    
    current_row += 1
    
    # Row 2: Total Credit & Advance Balance
    ws.cell(row=current_row, column=1, value="Total Credit:").font = font_bold
    ws.cell(row=current_row, column=1).border = double_bottom_border
    
    ws.cell(row=current_row, column=2).border = double_bottom_border
    ws.cell(row=current_row, column=3).border = double_bottom_border
    ws.cell(row=current_row, column=4).border = double_bottom_border
    
    tc_cell = ws.cell(row=current_row, column=6, value=float(total_credit)) # Credit is in col 6
    tc_cell.font = font_bold
    tc_cell.number_format = '"₹"#,##,##0.00'
    tc_cell.alignment = Alignment(horizontal="right")
    tc_cell.border = double_bottom_border
    
    # We place 'Advance Balance' label at col 5, but we can't because we need to align with Credit
    # Wait, in the headers:
    # 1: Date, 2: Particular, 3: Invoice Link, 4: Return Link, 5: Debit, 6: Credit, 7: Running Balance
    # So Total Debit is in col 5. Total Credit is in col 6.
    ws.cell(row=current_row-1, column=5).value = float(total_debit) # we set it already above in td_cell
    # For Advance balance label, we can just put it in col 5 for row 2, and value in col 7.
    
    ws.cell(row=current_row, column=5, value="Advance Balance:").font = font_bold
    ws.cell(row=current_row, column=5).border = double_bottom_border
    
    ab_cell = ws.cell(row=current_row, column=7, value=float(advance_balance))
    ab_cell.font = font_bold
    ab_cell.number_format = '"₹"#,##,##0.00'
    ab_cell.alignment = Alignment(horizontal="right")
    ab_cell.border = double_bottom_border
    ws.row_dimensions[current_row].height = 22
    
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Scan body rows only for width calculation
        for cell in col[8:current_row-1]: # exclude summary rows from width measurement to avoid weird spacing
            if cell.value:
                val_str = str(cell.value)
                if isinstance(cell.value, float) or isinstance(cell.value, Decimal):
                    val_str = f"₹{cell.value:,.2f}"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"ledger_{party.name.replace(' ', '_')}_{resolved_from.isoformat()}_to_{resolved_to.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/customer-ledger/{customer_id}")
async def get_customer_ledger_report(
    customer_id: uuid.UUID,
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    resolved_from, resolved_to = resolve_dates(from_date, to_date, month, year)
    
    party = await party_repository.get_by_id(db, customer_id)
    if not party:
        raise HTTPException(status_code=404, detail="Customer not found")
        
    ledger_data = await ledger_service.get_ledger(
        db, party_id=customer_id, from_date=resolved_from, to_date=resolved_to
    )
    
    entries = ledger_data.get("entries", [])
    opening_balance = Decimal(str(ledger_data.get("opening_balance", 0.0)))
    
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    for e in entries:
        total_debit += Decimal(str(e.get("debit", 0.0)))
        total_credit += Decimal(str(e.get("credit", 0.0)))
    
    final_balance = opening_balance + total_debit - total_credit
    current_outstanding = final_balance if final_balance > 0 else Decimal("0.00")
    advance_balance = abs(final_balance) if final_balance < 0 else Decimal("0.00")
    
    return {
        "party": {
            "name": party.name,
            "mobile": party.mobile or "",
            "address": party.address or "",
            "opening_balance": float(party.opening_balance),
            "balance": float(party.balance)
        },
        "from_date": resolved_from.isoformat(),
        "to_date": resolved_to.isoformat(),
        "opening_balance": float(opening_balance),
        "entries": entries,
        "summary": {
            "total_debit": float(total_debit),
            "total_credit": float(total_credit),
            "current_outstanding": float(current_outstanding),
            "advance_balance": float(advance_balance)
        }
    }
